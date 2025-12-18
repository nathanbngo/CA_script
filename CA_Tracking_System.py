"""
Corporate Action Tracking System
Processes daily CA files and creates Excel workbook with upcoming, recent, and archived CAs.
"""

import pandas as pd
import os
import glob
from datetime import datetime, timedelta, date
from pathlib import Path
import tkinter as tk
from tkinter import messagebox, filedialog
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# Configuration
INPUT_FOLDER = r"F:\Trade Support\Corporate Actions\CA check\CA Raw file"
# Folder where tracking workbooks and log will be stored
TRACKING_FOLDER = r"F:\Trade Support\Corporate Actions\CA check\CA Tracking"
MASTER_BASENAME = "CA_Tracking"

# Date ranges
TODAY = date.today()
NEXT_15_DAYS = TODAY + timedelta(days=15)
LAST_7_DAYS = TODAY - timedelta(days=7)

# Columns to extract (by name from CSV)
REQUIRED_COLUMNS = [
    "Security ID",
    "Security Name",
    "Event Type",
    "Response Status(ELIG)",
    "Client",
    "Reference ID",
    "Action Class",
    "ISIN",
    "Client Deadline Date",
    "Early Deadline Date"
]

# Filter criteria
EXCLUDED_EVENT_TYPES = [
    "OPTIONAL DIVIDEND",
    "CASH DISTRIBUTIONS",
    "DIVIDEND REINVESTMENT               "  # Note: includes trailing spaces
]


def print_progress(message):
    """Print progress message to console"""
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    line = f"[{timestamp}] {message}"
    print(line)


def show_message(title, message, is_error=False):
    """Show pop-up message box"""
    if is_error:
        messagebox.showerror(title, message)
    else:
        messagebox.showinfo(title, message)


def normalize_value(val):
    """Normalize values for comparison/logging (string, stripped; NaN -> empty)."""
    if pd.isna(val):
        return ""
    return str(val).strip()


def select_input_file(folder_path):
    """
    Determine which CA file to use for non-GUI runs.
    For GUI runs, the path is supplied directly.
    """
    latest_file = find_latest_file(folder_path)
    print_progress(f"Selected latest file automatically: {os.path.basename(latest_file)}")
    return latest_file


def run_ca_tracking(input_file=None, output_file=None, reset_archive=False):
    """Core execution logic, optionally with explicit input and output paths.

    If reset_archive is True, a brand-new tracking file is created using only
    the current input file (no previous archive is loaded).
    """
    print("=" * 60)
    print("Corporate Action Tracking System")
    print("=" * 60)
    print()

    # Step 1: Determine input file
    if input_file is None:
        input_file = select_input_file(INPUT_FOLDER)
    else:
        print_progress(f"Using input file provided by GUI: {os.path.basename(input_file)}")

    # Step 1b: Determine tracking folder and output file name
    # If output_file is provided by the GUI, treat it as a folder path.
    if output_file is None or output_file == "":
        tracking_folder = TRACKING_FOLDER
    else:
        tracking_folder = output_file

    # Ensure tracking folder exists
    os.makedirs(tracking_folder, exist_ok=True)

    # Create a new timestamped workbook path inside the tracking folder
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = os.path.join(tracking_folder, f"{MASTER_BASENAME}_{timestamp}.xlsx")

    # Step 2: Load data (ALL CAs, no filtering yet)
    df = load_data(input_file)

    # Step 3: Extract columns
    df = extract_columns(df)

    # Step 4: Parse dates
    df = parse_dates(df)

    # Step 5: Load existing Excel archive from latest tracking file (if any)
    latest_tracking_file = None
    previous_tab1_df = pd.DataFrame()

    if not reset_archive:
        try:
            pattern = os.path.join(tracking_folder, f"{MASTER_BASENAME}_*.xlsx")
            candidates = glob.glob(pattern)
            if candidates:
                latest_tracking_file = max(candidates, key=os.path.getmtime)
                print_progress(f"Latest tracking file detected: {os.path.basename(latest_tracking_file)}")
                try:
                    print_progress(f"Loading 'Next 15 Days' tab from previous file for comment transfer")
                    previous_tab1_df = pd.read_excel(latest_tracking_file, sheet_name="Next 15 Days", engine='openpyxl')
                    print_progress(f"Loaded {len(previous_tab1_df)} CAs from previous Next 15 Days tab")
                except Exception as e:
                    print_progress(f"Could not load previous Next 15 Days tab: {str(e)}")
                    previous_tab1_df = pd.DataFrame()
        except Exception:
            latest_tracking_file = None
            previous_tab1_df = pd.DataFrame()

        existing_archive_df, _ = load_existing_excel(latest_tracking_file)
    else:
        # Start fresh: no previous archive, no previous Tab 1 membership
        existing_archive_df = pd.DataFrame()
        previous_tab1_df = pd.DataFrame()

    # Step 6: Merge ALL CAs with archive (add new, update if changed, preserve comments)
    archive_df, change_summary = merge_with_archive(df, existing_archive_df)

    # Step 7: Generate Tab 1 and Tab 2 FROM archive (apply filters + date criteria)
    # Pass previous Tab 1 to transfer comments
    tab1_df, tab2_df = generate_tabs_from_archive(archive_df, previous_tab1_df)

    # Compute Next 15 Days membership changes (based on Reference ID)
    def refid_set(df):
        if df is None or df.empty:
            return set()
        return {
            str(x)
            for x in df.get("Reference ID", [])
            if pd.notna(x) and str(x) != ""
        }

    prev_tab1_ids = refid_set(previous_tab1_df)
    new_tab1_ids = refid_set(tab1_df)

    added_next15_ids = sorted(new_tab1_ids - prev_tab1_ids)
    removed_next15_ids = sorted(prev_tab1_ids - new_tab1_ids)

    changed_ids_for_tabs = set(change_summary.get("updated", [])) | set(
        change_summary.get("status_only", [])
    )
    updated_next15_ids = sorted(
        (new_tab1_ids & prev_tab1_ids) & changed_ids_for_tabs
    )

    # Step 8: Backup existing file (if it exists)
    backup_existing_file(file_path)

    # Step 9: Save to Excel
    save_to_excel(tab1_df, tab2_df, archive_df, file_path)

    # Step 10: Write per-run change log into Logs subfolder of tracking folder
    try:
        # Determine which changed CAs are urgent (<= 3 days) or in last 7 days
        added_ids = set(change_summary.get("added", []))
        updated_ids = set(change_summary.get("updated", []))
        status_only_ids = set(change_summary.get("status_only", []))
        changed_ids = added_ids.union(updated_ids).union(status_only_ids)

        urgent_window_end = TODAY + timedelta(days=3)

        urgent_changed = []
        last7_changed = []

        if not archive_df.empty and changed_ids:
            # Map Reference ID -> row for quick lookup
            index_by_ref = {}
            for _, row in archive_df.iterrows():
                ref_id = str(row.get("Reference ID", ""))
                if ref_id:
                    index_by_ref[ref_id] = row

            for ref_id in changed_ids:
                row = index_by_ref.get(ref_id)
                if row is None:
                    continue
                deadline = row.get("Deadline Date", None)
                if pd.isna(deadline) or deadline == "":
                    continue
                # Normalize deadline to a date object
                if isinstance(deadline, pd.Timestamp):
                    d = deadline.date()
                elif isinstance(deadline, datetime):
                    d = deadline.date()
                elif isinstance(deadline, date):
                    d = deadline
                else:
                    try:
                        d = pd.to_datetime(str(deadline)).date()
                    except Exception:
                        continue

                if TODAY <= d <= urgent_window_end:
                    urgent_changed.append((ref_id, row))
                elif LAST_7_DAYS <= d < TODAY:
                    last7_changed.append((ref_id, row))

        log_name = os.path.splitext(os.path.basename(file_path))[0] + ".log"
        log_folder = os.path.join(tracking_folder, "Logs")
        os.makedirs(log_folder, exist_ok=True)
        per_run_log_path = os.path.join(log_folder, log_name)
        with open(per_run_log_path, "w", encoding="utf-8") as f:
            f.write(f"CA Tracking Run Log - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Input file: {input_file}\n")
            f.write(f"Tracking workbook created: {file_path}\n")
            if latest_tracking_file:
                f.write(f"Previous tracking workbook: {latest_tracking_file}\n")
            else:
                f.write("Previous tracking workbook: None (first run)\n")
            f.write("\nSummary of changes:\n")
            f.write(f"  Added: {len(change_summary.get('added', []))}\n")
            f.write(f"  Updated (core fields): {len(change_summary.get('updated', []))}\n")
            f.write(f"  Status-only changes (Client / Response): {len(change_summary.get('status_only', []))}\n")
            f.write(f"  Unchanged: {len(change_summary.get('unchanged', []))}\n")
            f.write(f"  In archive but missing from latest input: {len(change_summary.get('missing_from_input', []))}\n")
            f.write(f"  Urgent (<= 3 days) added/updated: {len(urgent_changed)}\n")
            f.write(f"  Last 7 days added/updated: {len(last7_changed)}\n")
            f.write(
                f"  Next 15 Days - Added: {len(added_next15_ids)}, "
                f"Updated: {len(updated_next15_ids)}, "
                f"Removed: {len(removed_next15_ids)}\n"
            )

            def write_id_list(title, items):
                if not items:
                    return
                f.write(f"\n{title} ({len(items)} IDs):\n")
                for ref_id in items:
                    f.write(f"  - {ref_id}\n")

            write_id_list("Added Reference IDs", change_summary.get("added", []))
            write_id_list("Updated Reference IDs", change_summary.get("updated", []))
            # We intentionally do NOT list status-only IDs per RefID to keep logs focused
            write_id_list(
                "Reference IDs in archive but missing from latest input",
                change_summary.get("missing_from_input", []),
            )
            write_id_list("Next 15 Days - Added Reference IDs", added_next15_ids)
            write_id_list("Next 15 Days - Updated Reference IDs", updated_next15_ids)
            write_id_list("Next 15 Days - Removed Reference IDs", removed_next15_ids)

            def write_ca_detail_list(title, items):
                if not items:
                    return
                f.write(f"\n{title} ({len(items)}):\n")
                for ref_id, row in items:
                    security_name = str(row.get("Security Name", ""))
                    event_type = str(row.get("Event Type", ""))
                    deadline = row.get("Deadline Date", "")
                    status = "Added" if ref_id in added_ids else "Updated"
                    f.write(f"  - RefID: {ref_id}\n")
                    f.write(f"    Status   : {status}\n")
                    f.write(f"    Security : {security_name}\n")
                    f.write(f"    Event    : {event_type}\n")
                    f.write(f"    Deadline : {deadline}\n")
                    if status == "Updated":
                        details = change_summary.get("updated_details", {}).get(ref_id)
                        if details and details.get("changed_columns"):
                            f.write(f"    Changed fields:\n")
                            for col in details["changed_columns"]:
                                old_val = details["old_values"].get(col, "")
                                new_val = details["new_values"].get(col, "")
                                f.write(f"      - {col}: \"{old_val}\" -> \"{new_val}\"\n")
                    f.write("\n")

            write_ca_detail_list("Urgent CAs (<= 3 days) added/updated", urgent_changed)
            write_ca_detail_list("Last 7 days CAs added/updated", last7_changed)

            # General details for all updated CAs (regardless of deadline bucket)
            updated_details_map = change_summary.get("updated_details", {})

            if updated_details_map:
                f.write("\nUpdated CAs (core fields only, full details):\n")
                for ref_id in change_summary.get("updated", []):
                    # Find row in archive
                    row = archive_df[archive_df["Reference ID"].astype(str) == str(ref_id)]
                    row_data = row.iloc[0] if not row.empty else {}
                    security_name = str(row_data.get("Security Name", "")) if isinstance(row_data, dict) else str(getattr(row_data, "get", lambda *_: "")("Security Name", ""))
                    event_type = str(row_data.get("Event Type", "")) if isinstance(row_data, dict) else str(getattr(row_data, "get", lambda *_: "")("Event Type", ""))
                    deadline = row_data.get("Deadline Date", "") if isinstance(row_data, dict) else getattr(row_data, "get", lambda *_: "")("Deadline Date", "")

                    f.write(f"\n  RefID: {ref_id}\n")
                    f.write(f"    Security : {security_name}\n")
                    f.write(f"    Event    : {event_type}\n")
                    f.write(f"    Deadline : {deadline}\n")

                    details = updated_details_map.get(ref_id)
                    if details and details.get("changed_columns"):
                        f.write(f"    Changed fields:\n")
                        for col in details["changed_columns"]:
                            old_val = details["old_values"].get(col, "")
                            new_val = details["new_values"].get(col, "")
                            f.write(f"      - {col}: \"{old_val}\" -> \"{new_val}\"\n")
    except Exception as e:
        print_progress(f"Warning: failed to write per-run log: {e}")

    # Success message
    print()
    print("=" * 60)
    print("SUCCESS!")
    print("=" * 60)
    print(f"Tab 1 (Next 15 Days): {len(tab1_df)} CAs")
    print(f"Tab 2 (Last 7 Days): {len(tab2_df)} CAs")
    print(f"Archive: {len(archive_df)} CAs")
    print(f"Output file: {file_path}")
    print()

    show_message(
        "Success",
        f"CA Tracking updated successfully!\n\n"
        f"Next 15 Days: {len(tab1_df)} CAs\n"
        f"Last 7 Days: {len(tab2_df)} CAs\n"
        f"Archive: {len(archive_df)} CAs\n\n"
        f"File saved to:\n{file_path}"
    )


def find_latest_file(folder_path):
    """Find the latest CA file (CSV or Excel) in the folder"""
    print_progress(f"Searching for CSV/Excel files in: {folder_path}")
    
    if not os.path.exists(folder_path):
        raise FileNotFoundError(f"Input folder not found: {folder_path}")
    
    # Look for supported files
    csv_files = glob.glob(os.path.join(folder_path, "*.csv"))
    xls_files = glob.glob(os.path.join(folder_path, "*.xls"))
    xlsx_files = glob.glob(os.path.join(folder_path, "*.xlsx"))
    
    all_files = csv_files + xls_files + xlsx_files
    
    if not all_files:
        raise FileNotFoundError(f"No CSV or Excel files found in: {folder_path}")
    
    # Get the most recently modified file
    latest_file = max(all_files, key=os.path.getmtime)
    print_progress(f"Found latest file: {os.path.basename(latest_file)}")
    return latest_file


def load_data(file_path):
    """Load data from CA file (CSV or Excel)"""
    print_progress(f"Loading data from: {os.path.basename(file_path)}")
    
    file_ext = os.path.splitext(file_path)[1].lower()
    
    if file_ext == '.csv':
        df = pd.read_csv(file_path)
    elif file_ext in ['.xls', '.xlsx']:
        df = pd.read_excel(file_path)
    else:
        raise ValueError(f"Unsupported file format: {file_ext}. Expected .csv, .xls or .xlsx")
    
    print_progress(f"Loaded {len(df)} rows")
    return df


def extract_columns(df):
    """Extract required columns from dataframe"""
    print_progress("Extracting required columns...")
    
    # Check which columns exist
    missing_cols = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing_cols:
        raise ValueError(f"Missing required columns: {missing_cols}")
    
    # Extract columns
    output = df[REQUIRED_COLUMNS].copy()

    # Bring through Comments from input file if present (optional)
    if "Comments" in df.columns:
        output["Comments"] = df["Comments"]
    
    # Remove rows where all key columns are empty
    output = output.dropna(subset=["Security ID", "Security Name"], how='all')
    
    return output


def parse_dates(df):
    """Parse date columns"""
    print_progress("Parsing date columns...")
    
    # Parse Client Deadline Date
    if 'Client Deadline Date' in df.columns:
        # Handle date format: "29 Dec 2050 03:30:00 PM EST" -> take first 11 chars
        df['Client Deadline Date'] = df['Client Deadline Date'].astype(str)
        df['Client Deadline Date'] = pd.to_datetime(
            df['Client Deadline Date'].str[:11], 
            format='%d %b %Y', 
            errors='coerce'
        )
    
    # Parse Early Deadline Date
    if 'Early Deadline Date' in df.columns:
        df['Early Deadline Date'] = df['Early Deadline Date'].astype(str)
        df['Early Deadline Date'] = pd.to_datetime(
            df['Early Deadline Date'].str[:11], 
            format='%d %b %Y', 
            errors='coerce'
        )
    
    return df


def apply_filters(df):
    """Apply all filter criteria"""
    print_progress("Applying filters...")
    
    initial_count = len(df)
    
    # Filter 1: Response Status – exclude "NOT APPLICABLE"
    df = df[~df['Response Status(ELIG)'].str.contains("NOT APPLICABLE", na=False)]
    
    # Filter 2: Client not equal to "nil" or empty
    df = df[df['Client'].notna()]
    df = df[df['Client'] != "nil"]
    df = df[df['Client'] != ""]
    
    # Filter 3: Event Type NOT IN excluded list
    df = df[~df['Event Type'].isin(EXCLUDED_EVENT_TYPES)]
    
    # Filter 4: Action Class not equal to "Mandatory"
    df = df[df['Action Class'] != "Mandatory"]
    
    filtered_count = len(df)
    print_progress(f"Filtered from {initial_count} to {filtered_count} rows")
    
    return df


def determine_deadline_date(row):
    """
    Determine which deadline date to use and return (date, type)
    Logic:
    - If both dates are in next 15 days: use earlier date
    - If Early Deadline is in next 15 days: use Early Deadline
    - If Client Deadline is in next 15 days: use Client Deadline
    - If Early Deadline is past but Client Deadline is future: use Client Deadline
    - If Client Deadline is in last 7 days: use Client Deadline
    """
    client_deadline = row['Client Deadline Date']
    early_deadline = row['Early Deadline Date']
    
    # Convert to date if datetime
    if pd.notna(client_deadline):
        if isinstance(client_deadline, pd.Timestamp):
            client_deadline = client_deadline.date()
        elif isinstance(client_deadline, datetime):
            client_deadline = client_deadline.date()
    
    if pd.notna(early_deadline):
        if isinstance(early_deadline, pd.Timestamp):
            early_deadline = early_deadline.date()
        elif isinstance(early_deadline, datetime):
            early_deadline = early_deadline.date()
    
    # Check if dates are in range
    client_in_next_15 = (pd.notna(client_deadline) and 
                        TODAY <= client_deadline <= NEXT_15_DAYS)
    early_in_next_15 = (pd.notna(early_deadline) and 
                       TODAY <= early_deadline <= NEXT_15_DAYS)
    
    client_in_last_7 = (pd.notna(client_deadline) and 
                       LAST_7_DAYS <= client_deadline < TODAY)
    early_in_last_7 = (pd.notna(early_deadline) and 
                      LAST_7_DAYS <= early_deadline < TODAY)
    
    # Priority: Use earlier date if both are in next 15 days
    if client_in_next_15 and early_in_next_15:
        if early_deadline <= client_deadline:
            return early_deadline, "Early"
        else:
            return client_deadline, "Client"
    
    # Use Early Deadline if in next 15 days
    if early_in_next_15:
        return early_deadline, "Early"
    
    # Use Client Deadline if in next 15 days
    if client_in_next_15:
        return client_deadline, "Client"
    
    # Check for last 7 days (Tab 2)
    if early_in_last_7:
        return early_deadline, "Early"
    
    if client_in_last_7:
        return client_deadline, "Client"
    
    # If Early Deadline is past but Client Deadline is future (beyond 15 days)
    if (pd.notna(early_deadline) and early_deadline < TODAY and 
        pd.notna(client_deadline) and client_deadline > NEXT_15_DAYS):
        return client_deadline, "Client"
    
    # No valid date in range
    return None, None


def generate_tabs_from_archive(archive_df, previous_tab1_df=None):
    """Generate Tab 1 and Tab 2 from archive by applying filters and date criteria"""
    print_progress("Generating tabs from archive...")
    
    if archive_df.empty:
        return pd.DataFrame(), pd.DataFrame()
    
    # First, apply filters to archive (same filters as before)
    filtered_df = apply_filters_to_archive(archive_df)
    
    # Create dictionary of previous Tab 1 comments by Reference ID
    # Only include CAs that are still within 15 days of current date
    previous_comments = {}
    if previous_tab1_df is not None and not previous_tab1_df.empty:
        if 'Reference ID' in previous_tab1_df.columns and 'Comments' in previous_tab1_df.columns and 'Deadline Date' in previous_tab1_df.columns:
            for idx, row in previous_tab1_df.iterrows():
                ref_id = str(row.get('Reference ID', '')).strip()
                comment = row.get('Comments', '')
                deadline_date = row.get('Deadline Date', None)
                
                # Parse deadline date and check if still within 15 days
                if pd.notna(deadline_date) and deadline_date != "":
                    try:
                        if isinstance(deadline_date, str):
                            deadline_date = pd.to_datetime(deadline_date).date()
                        elif isinstance(deadline_date, pd.Timestamp):
                            deadline_date = deadline_date.date()
                        elif isinstance(deadline_date, datetime):
                            deadline_date = deadline_date.date()
                        
                        # Only include if still within next 15 days
                        if pd.notna(ref_id) and ref_id != '' and TODAY <= deadline_date <= NEXT_15_DAYS:
                            comment_str = str(comment).strip() if pd.notna(comment) and str(comment).strip() != '' else ''
                            if comment_str:  # Only store non-empty comments
                                previous_comments[ref_id] = comment_str
                    except:
                        continue
            
            print_progress(f"Loaded {len(previous_comments)} comments from previous Next 15 Days tab (still within date range)")
    
    tab1_data = []
    tab2_data = []
    
    for idx, row in filtered_df.iterrows():
        deadline_date = row.get('Deadline Date', None)
        
        if pd.isna(deadline_date) or deadline_date == "":
            continue
        
        # Convert to date if needed
        if isinstance(deadline_date, str):
            try:
                deadline_date = pd.to_datetime(deadline_date).date()
            except:
                continue
        elif isinstance(deadline_date, pd.Timestamp):
            deadline_date = deadline_date.date()
        elif isinstance(deadline_date, datetime):
            deadline_date = deadline_date.date()
        
        row_data = row.copy()
        
        # Tab 1: Next 15 days
        if TODAY <= deadline_date <= NEXT_15_DAYS:
            ref_id = str(row_data.get('Reference ID', '')).strip()
            today_comment = row_data.get('Comments', '')
            
            # Priority: Today's comment (if exists) overrides previous comment
            # If no today's comment, transfer from previous tab if available
            has_today_comment = (pd.notna(today_comment) and str(today_comment).strip() != '')
            
            if has_today_comment:
                # Use today's comment (override)
                row_data['Comments'] = str(today_comment).strip()
            elif ref_id in previous_comments:
                # No today's comment, but previous tab had one - transfer it
                row_data['Comments'] = previous_comments[ref_id]
            else:
                # No comment from either source
                row_data['Comments'] = ""
            
            tab1_data.append(row_data)
        # Tab 2: Last 7 days (auto-remove after 7 days)
        elif LAST_7_DAYS <= deadline_date < TODAY:
            tab2_data.append(row_data)
    
    tab1_df = pd.DataFrame(tab1_data) if tab1_data else pd.DataFrame()
    tab2_df = pd.DataFrame(tab2_data) if tab2_data else pd.DataFrame()
    
    # Sort by deadline date
    if not tab1_df.empty and 'Deadline Date' in tab1_df.columns:
        tab1_df = tab1_df.sort_values(by='Deadline Date')
    if not tab2_df.empty and 'Deadline Date' in tab2_df.columns:
        tab2_df = tab2_df.sort_values(by='Deadline Date')
    
    # Count transferred comments
    if not tab1_df.empty:
        comments_transferred = sum(1 for idx, row in tab1_df.iterrows() 
                                 if str(row.get('Comments', '')).strip() != '' 
                                 and str(row.get('Reference ID', '')).strip() in previous_comments
                                 and str(row.get('Comments', '')).strip() == previous_comments.get(str(row.get('Reference ID', '')).strip(), ''))
        if comments_transferred > 0:
            print_progress(f"Transferred {comments_transferred} comments from previous Next 15 Days tab")
    
    print_progress(f"Tab 1 (Next 15 Days): {len(tab1_df)} CAs")
    print_progress(f"Tab 2 (Last 7 Days): {len(tab2_df)} CAs")
    
    return tab1_df, tab2_df


def apply_filters_to_archive(df):
    """Apply filter criteria to archive data (for Tab 1 and Tab 2)"""
    # Filter 1: Response Status – exclude "NOT APPLICABLE"
    df = df[~df['Response Status(ELIG)'].str.contains("NOT APPLICABLE", na=False)]
    
    # Filter 2: Client not equal to "nil" or empty
    df = df[df['Client'].notna()]
    df = df[df['Client'] != "nil"]
    df = df[df['Client'] != ""]
    
    # Filter 3: Event Type NOT IN excluded list
    df = df[~df['Event Type'].isin(EXCLUDED_EVENT_TYPES)]
    
    # Filter 4: Action Class not equal to "Mandatory"
    df = df[df['Action Class'] != "Mandatory"]
    
    return df


def prepare_output_columns(df):
    """Prepare columns for output"""
    output_cols = [
        "Security ID",
        "Security Name",
        "Event Type",
        "Response Status(ELIG)",
        "Client",
        "Reference ID",
        "Action Class",
        "ISIN",
        "Deadline Date",
        "Deadline Type",
        "Comments"
    ]
    
    # Ensure all columns exist
    for col in output_cols:
        if col not in df.columns:
            df[col] = ""
    
    # Select and reorder columns
    return df[output_cols]


def load_existing_excel(file_path):
    """Load existing Excel file if it exists"""
    if not file_path or not os.path.exists(file_path):
        return pd.DataFrame(), pd.DataFrame()
    
    print_progress(f"Loading existing Excel file: {os.path.basename(file_path)}")
    
    try:
        # Load archive tab
        archive_df = pd.read_excel(file_path, sheet_name="Archive", engine='openpyxl')
        print_progress(f"Loaded {len(archive_df)} CAs from archive")
    except Exception:
        archive_df = pd.DataFrame()
    
    try:
        # Load Tab 2 to preserve manual removals
        tab2_df = pd.read_excel(file_path, sheet_name="Last 7 Days", engine='openpyxl')
        print_progress(f"Loaded {len(tab2_df)} CAs from Tab 2")
    except Exception:
        tab2_df = pd.DataFrame()
    
    return archive_df, tab2_df


def data_changed(
    new_row,
    existing_row,
    exclude_cols=None,
):
    """Check if data has changed (excluding specified columns)"""
    if exclude_cols is None:
        # Also ignore Client and Response Status(ELIG) to avoid noisy updates
        exclude_cols = ['Comments', 'Deadline Date', 'Deadline Type', 'Client', 'Response Status(ELIG)']

    # Only compare columns that exist in both rows and are not excluded
    for col in new_row.index:
        if col in exclude_cols:
            continue
        if col not in existing_row.index:
            # Column doesn't exist in archive row (e.g., raw deadline columns not stored) -> ignore
            continue

        new_val = normalize_value(new_row[col])
        existing_val = normalize_value(existing_row.get(col, None))

        if new_val != existing_val:
            return True
    return False


def merge_with_archive(new_df, existing_archive_df):
    """Merge new data with archive, preserving comments and only updating if data changed.
    Returns the updated archive and a change summary (counts + reference ID lists).
    """
    print_progress("Merging with archive...")

    change_summary = {
        "added": [],
        "updated": [],
        "unchanged": [],
        "missing_from_input": [],  # present in archive but not in latest Nexen file
        "status_only": [],         # only Client / Response Status changed
        "updated_details": {},     # ref_id -> {changed_columns, old_values, new_values}
        "status_details": {},      # ref_id -> {changed_columns, old_values, new_values}
    }

    # First time: no existing archive
    if existing_archive_df.empty:
        new_df = new_df.copy()
        # Initialize Comments: use input Comments if present, else empty
        if "Comments" in new_df.columns:
            new_df["Comments"] = new_df["Comments"].fillna("")
        else:
            new_df["Comments"] = ""
        deadline_dates = []
        deadline_types = []
        for idx, row in new_df.iterrows():
            deadline_date, deadline_type = determine_deadline_date(row)
            deadline_dates.append(deadline_date if deadline_date else "")
            deadline_types.append(deadline_type if deadline_type else "")
            ref_id = str(row.get("Reference ID", ""))
            if ref_id:
                change_summary["added"].append(ref_id)
        new_df["Deadline Date"] = deadline_dates
        new_df["Deadline Type"] = deadline_types

        print_progress(f"Archive: {len(change_summary['added'])} added, 0 updated, 0 unchanged")
        print_progress(f"Archive now contains {len(new_df)} CAs")
        return new_df, change_summary

    # Create archive copy
    archive_df = existing_archive_df.copy()

    # Compute missing_from_input: IDs in archive that are not in latest Nexen file
    try:
        existing_ids = set(str(x) for x in archive_df.get("Reference ID", []) if pd.notna(x) and str(x) != "")
        new_ids = set(str(x) for x in new_df.get("Reference ID", []) if pd.notna(x) and str(x) != "")
        missing_ids = sorted(existing_ids - new_ids)
        change_summary["missing_from_input"] = missing_ids
        if missing_ids:
            print_progress(f"{len(missing_ids)} CAs are in archive but missing from latest input (kept in archive).")
    except Exception:
        pass

    # Create a dictionary of existing rows by Reference ID
    existing_dict = {}
    if "Reference ID" in archive_df.columns:
        for idx, row in archive_df.iterrows():
            ref_id = str(row.get("Reference ID", ""))
            if pd.notna(ref_id) and ref_id != "":
                existing_dict[ref_id] = (idx, row)

    # Process new data
    added_count = 0
    updated_count = 0
    skipped_count = 0
    status_only_count = 0

    for idx, row in new_df.iterrows():
        ref_id = str(row.get("Reference ID", ""))

        if not ref_id or ref_id == "":
            continue

        # Calculate deadline date and type for new row
        deadline_date, deadline_type = determine_deadline_date(row)
        row["Deadline Date"] = deadline_date if deadline_date else ""
        row["Deadline Type"] = deadline_type if deadline_type else ""

        if ref_id in existing_dict:
            # Reference ID exists - check if data changed
            existing_idx, existing_row = existing_dict[ref_id]

            # Determine which fields changed
            core_changed_columns = []
            status_changed_columns = []
            core_old_values = {}
            core_new_values = {}
            status_old_values = {}
            status_new_values = {}

            STATUS_COLS = ["Client", "Response Status(ELIG)"]

            for col in row.index:
                if col in ["Comments", "Deadline Date", "Deadline Type"]:
                    continue
                if col not in existing_row.index:
                    continue

                new_val = normalize_value(row[col])
                old_val = normalize_value(existing_row.get(col, None))

                if new_val != old_val:
                    if col in STATUS_COLS:
                        status_changed_columns.append(col)
                        status_old_values[col] = old_val
                        status_new_values[col] = new_val
                    else:
                        core_changed_columns.append(col)
                        core_old_values[col] = old_val
                        core_new_values[col] = new_val

            if core_changed_columns or status_changed_columns:
                # Some change detected -> update archive row fields
                for col in row.index:
                    if col == "Comments":
                        continue
                    archive_df.at[existing_idx, col] = row[col]
                archive_df.at[existing_idx, "Deadline Date"] = deadline_date if deadline_date else ""
                archive_df.at[existing_idx, "Deadline Type"] = deadline_type if deadline_type else ""

                # Handle Comments: overwrite only if new comment is non-blank,
                # otherwise preserve existing archive comment
                if "Comments" in row.index:
                    new_comment = normalize_value(row["Comments"])
                    if new_comment != "":
                        archive_df.at[existing_idx, "Comments"] = row["Comments"]

                if core_changed_columns:
                    updated_count += 1
                    change_summary["updated"].append(ref_id)
                    change_summary["updated_details"][ref_id] = {
                        "changed_columns": core_changed_columns,
                        "old_values": core_old_values,
                        "new_values": core_new_values,
                    }

                elif status_changed_columns:
                    status_only_count += 1
                    change_summary["status_only"].append(ref_id)
                    change_summary["status_details"][ref_id] = {
                        "changed_columns": status_changed_columns,
                        "old_values": status_old_values,
                        "new_values": status_new_values,
                    }
            else:
                # No changes - only ensure deadline fields are normalized
                archive_df.at[existing_idx, "Deadline Date"] = deadline_date if deadline_date else ""
                archive_df.at[existing_idx, "Deadline Type"] = deadline_type if deadline_type else ""
                skipped_count += 1
                change_summary["unchanged"].append(ref_id)

            # Comments handling: overwrite only if new comment is non-blank.
            # This applies even when only the comment changed.
            if "Comments" in row.index:
                new_comment = normalize_value(row["Comments"])
                if new_comment != "":
                    existing_comment = ""
                    if "Comments" in existing_row.index:
                        existing_comment = normalize_value(existing_row["Comments"])
                    if new_comment != existing_comment:
                        archive_df.at[existing_idx, "Comments"] = row["Comments"]
        else:
            # New CA - add to archive with comment from input if present
            new_row = row.copy()
            if "Comments" in new_row.index:
                new_comment = normalize_value(new_row["Comments"])
                new_row["Comments"] = new_comment
            else:
                new_row["Comments"] = ""
            archive_df = pd.concat([archive_df, pd.DataFrame([new_row])], ignore_index=True)
            added_count += 1
            change_summary["added"].append(ref_id)

    print_progress(
        f"Archive: {added_count} added, {updated_count} updated, "
        f"{status_only_count} status-only changes, {skipped_count} unchanged"
    )
    print_progress(f"Archive now contains {len(archive_df)} CAs")
    return archive_df, change_summary


def format_excel(file_path):
    """Apply formatting to Excel file"""
    print_progress("Applying Excel formatting...")
    
    wb = openpyxl.load_workbook(file_path)
    
    for sheet_name in ["Next 15 Days", "Last 7 Days", "Archive"]:
        if sheet_name not in wb.sheetnames:
            continue
        
        ws = wb[sheet_name]
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Format header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Conditional formatting for Tab 1 (Next 15 Days)
        if sheet_name == "Next 15 Days":
            # Find Deadline Date column
            deadline_col = None
            for idx, cell in enumerate(ws[1], 1):
                if cell.value == "Deadline Date":
                    deadline_col = get_column_letter(idx)
                    break
            
            if deadline_col:
                # Red for < 3 days (urgent)
                red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                ws.conditional_formatting.add(
                    f"{deadline_col}2:{deadline_col}{ws.max_row}",
                    CellIsRule(operator='lessThan', formula=[f'TODAY()+3'], fill=red_fill)
                )
                
                # Yellow for 3-7 days (approaching)
                yellow_fill = PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid")
                ws.conditional_formatting.add(
                    f"{deadline_col}2:{deadline_col}{ws.max_row}",
                    CellIsRule(operator='between', formula=[f'TODAY()+3', f'TODAY()+7'], fill=yellow_fill)
                )
                
                # Green for > 7 days (upcoming)
                green_fill = PatternFill(start_color="95E1D3", end_color="95E1D3", fill_type="solid")
                ws.conditional_formatting.add(
                    f"{deadline_col}2:{deadline_col}{ws.max_row}",
                    CellIsRule(operator='greaterThan', formula=[f'TODAY()+7'], fill=green_fill)
                )
    
    wb.save(file_path)
    print_progress("Formatting complete")


def backup_existing_file(file_path):
    """Create backup of existing file"""
    if not os.path.exists(file_path):
        return
    
    print_progress("Creating backup...")
    
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    backup_path = file_path.replace(".xlsx", f"_backup_{timestamp}.xlsx")
    
    import shutil
    shutil.copy2(file_path, backup_path)
    print_progress(f"Backup created: {os.path.basename(backup_path)}")


def save_to_excel(tab1_df, tab2_df, archive_df, file_path):
    """Save dataframes to Excel with multiple tabs"""
    print_progress("Saving to Excel...")
    
    # Prepare dataframes
    if not tab1_df.empty:
        tab1_df = prepare_output_columns(tab1_df)
        tab1_df = tab1_df.sort_values(by='Deadline Date')
    else:
        tab1_df = pd.DataFrame(columns=[
            "Security ID", "Security Name", "Event Type", "Response Status(ELIG)",
            "Client", "Reference ID", "Action Class", "ISIN", "Deadline Date",
            "Deadline Type", "Comments"
        ])
    
    if not tab2_df.empty:
        tab2_df = prepare_output_columns(tab2_df)
        tab2_df = tab2_df.sort_values(by='Deadline Date')
    else:
        tab2_df = pd.DataFrame(columns=[
            "Security ID", "Security Name", "Event Type", "Response Status(ELIG)",
            "Client", "Reference ID", "Action Class", "ISIN", "Deadline Date",
            "Deadline Type", "Comments"
        ])
    
    if not archive_df.empty:
        archive_df = prepare_output_columns(archive_df)
    else:
        archive_df = pd.DataFrame(columns=[
            "Security ID", "Security Name", "Event Type", "Response Status(ELIG)",
            "Client", "Reference ID", "Action Class", "ISIN", "Deadline Date",
            "Deadline Type", "Comments"
        ])
    
    # Format dates for Excel
    for df in [tab1_df, tab2_df, archive_df]:
        if 'Deadline Date' in df.columns:
            df['Deadline Date'] = pd.to_datetime(df['Deadline Date']).dt.strftime('%Y-%m-%d')
    
    # Save to Excel
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        tab1_df.to_excel(writer, sheet_name="Next 15 Days", index=False)
        tab2_df.to_excel(writer, sheet_name="Last 7 Days", index=False)
        archive_df.to_excel(writer, sheet_name="Archive", index=False)
    
    # Apply formatting
    format_excel(file_path)
    
    print_progress(f"Excel file saved: {file_path}")


def main():
    """Main execution function (CLI/non-GUI entry point)"""
    try:
        run_ca_tracking()
    except Exception as e:
        error_msg = f"Error: {str(e)}"
        print_progress(error_msg)
        print()
        import traceback
        traceback.print_exc()
        show_message("Error", error_msg, is_error=True)


if __name__ == "__main__":
    # Initialize tkinter root (hidden) for message boxes
    root = tk.Tk()
    root.withdraw()
    
    main()
    
    root.destroy()

